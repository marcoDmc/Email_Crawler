from fastapi import FastAPI, HTTPException
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os

app = FastAPI()

# Load environment variables
load_dotenv()

# Configurações do Chrome
options = Options()
options.add_argument("--start-maximized")
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("--disable-extensions")
options.add_argument("--disable-popup-blocking")
options.add_argument("--disable-infobars")

# Inicialização do WebDriver
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)


# Função para abrir o navegador e acessar o Gmail
def open_browser(url):
    driver.get(url)


# Função para realizar o login no Gmail
@app.get()
def login():
    try:
        open_browser("https://www.google.com/intl/pt-BR/gmail/about/")  # URL do Gmail

        # Clicar no botão de login
        login_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/header/div/div/div/a[2]"))
        )
        login_button.click()

        # Preencher o email e avançar para a próxima tela
        email_input = WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.TAG_NAME, 'input'))
        )
        email_input.send_keys(os.getenv("EMAIL"))

        button_next_email = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="identifierNext"]/div/button'))
        )
        button_next_email.click()

        # Preencher a senha e avançar
        password_input = WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input'))
        )
        if password_input:
            password_input.send_keys(os.getenv('PASSWORD'))

        button_next_password = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="passwordNext"]/div/button'))
        )
        button_next_password.click()

        # Esperar a caixa de entrada do Gmail carregar
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'div[role="main"]'))
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error during login: {str(e)}")


# Função para extrair textos dos emails
def extract_all_emails_text():
    try:
        # Encontrar todos os emails na caixa de entrada
        email_rows = WebDriverWait(driver, 30).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'tr.zA'))
        )
        print(len(email_rows), 'EMAILS')

        emails_data = []

        for row in email_rows:
            try:
                # Capturar o texto do assunto do email
                subject_element = row.find_element(By.CSS_SELECTOR, 'span.bog')
                subject_text = subject_element.text

                # Capturar o texto do resumo do email
                snippet_element = row.find_element(By.CSS_SELECTOR, 'span.y2')
                snippet_text = snippet_element.text

                # Adicionar os dados do email ao dicionário
                emails_data.append({
                    "Assunto": subject_text,
                    "Resumo": snippet_text
                })

            except Exception as e:
                print(f"Erro ao extrair texto do email: {str(e)}")

        if len(emails_data) > 0:
            # Criar um novo arquivo Excel e configurar a formatação
            wb = Workbook()
            ws = wb.active
            ws.title = "Emails"

            # Configuração da formatação das células
            header_font = Font(bold=True, color="FFFFFF")
            fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")

            # Adicionar cabeçalho formatado
            ws['A1'] = "Assunto"
            ws['A1'].font = header_font
            ws['A1'].fill = fill

            ws['B1'] = "Resumo"
            ws['B1'].font = header_font
            ws['B1'].fill = fill

            # Adicionar dados dos emails
            row_num = 2
            for email in emails_data:
                ws[f'A{row_num}'] = email["Assunto"]
                ws[f'B{row_num}'] = email["Resumo"]
                row_num += 1

            # Ajustar largura das colunas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width

            # Salvar o arquivo Excel
            wb.save("emails_formatted.xlsx")
            print("Arquivo Excel 'emails_formatted.xlsx' criado com sucesso.")

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao tentar extrair textos dos emails: {str(e)}")


# Função para fechar o navegador
def close_browser():
    driver.quit()


# @app.get("/extract-emails/")
async def extract_emails():
    try:
        # login()
        extract_all_emails_text()
        close_browser()
        return {"message": "Emails extracted and saved successfully"}
    except Exception as e:
        close_browser()
        # raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
